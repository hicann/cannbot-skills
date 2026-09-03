#!/usr/bin/env python3
# -----------------------------------------------------------------------------------------------------------
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
# -----------------------------------------------------------------------------------------------------------

"""
从批跑输出目录生成结果 Excel，含算子精度/性能数据与运行元信息
（NPU/耗时/最终状态）。

数据来源优先级（每个算子独立判定）：
  1. {op_dir}/report.md           —— 优先解析
  2. {op_dir}/summary.json        —— report.md 缺失或字段缺失时回退
  3. {op_dir}/output/<迭代目录>     —— report.md 与 summary.json 均缺失时，扫描各迭代目录的
                                    *_perf_result*.json，取加速比最高版本及对应 verify 精度
  4. {op_dir}/output/perf_result.json —— 延迟字段补充（report/summary 存在时）
  5. {op_dir}/precheck.json       —— 提取 category/op_subtype

输出: {output_dir}/batch_report.xlsx（已存在时生成带时间戳的新文件，不覆盖）
"""

import argparse
import json
import logging
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _log_utils import setup_logger  # noqa: E402

logger = logging.getLogger("gen_batch_excel")


# ── batch_report.md 解析 ────────────────────────────────────────────────────
def parse_batch_report(path: Path):
    """返回 (header_info, op_rows)。
    header_info: dict —— benchmark / level / arch / npu / npu_list / mode / start_time / end_time
    op_rows: list of dict {id, file, status, elapsed}
    """
    text = path.read_text(encoding="utf-8")
    header = {}
    for key, pat in [
        ("benchmark", r"^- benchmark:\s*(.+)$"),
        ("level", r"^- level:\s*(.+)$"),
        ("arch", r"^- arch:\s*(.+)$"),
        ("npu", r"^- npu:\s*(.+)$"),
        ("npu_list", r"^- npu-list:\s*(.+)$"),
        ("mode", r"^- 执行模式:\s*(.+)$"),
        ("start_time", r"^- 开始时间:\s*(.+)$"),
        ("end_time", r"^- 结束时间:\s*(.+)$"),
    ]:
        m = re.search(pat, text, flags=re.MULTILINE)
        if m:
            header[key] = m.group(1).strip()

    op_rows = []
    for line in text.splitlines():
        m = re.match(
            r"^\|\s*(\d+)\s*\|\s*([^|]+\.py)\s*\|\s*([^|]+?)\s*\|\s*(\d+)\s*\|",
            line,
        )
        if m:
            op_rows.append({
                "id": int(m.group(1)),
                "file": m.group(2).strip(),
                "status": m.group(3).strip(),
                "elapsed": int(m.group(4)),
            })
    return header, op_rows


# ── report.md 解析 ──────────────────────────────────────────────────────────
REPORT_PATTERNS = {
    "op_name": r"\*\*算子名称\*\*:\s*(\S+)",
    "arch": r"\*\*架构\*\*:\s*(\S+)",
    "npu": r"\*\*NPU\*\*:\s*(\S+)",
    "target_speedup": r"\*\*目标加速比\*\*:\s*([\d.]+)",
    "target_reached": r"\*\*是否达到目标\*\*:\s*(\S+)",
    "best_speedup": r"\*\*实际最佳加速比\*\*:\s*([\d.]+)",
    "phase3_speedup": r"\*\*Phase 3 基线加速比\*\*:\s*([\d.]+)",
    "pass_rate_passed": r"\*\*通过率\*\*:\s*(\d+)\s*/\s*(\d+)",
    "fail_count": r"\*\*失败数\*\*:\s*(\d+)",
    "framework_latency_ms": r"\*\*框架平均延迟\*\*:\s*([\d.]+)\s*ms",
    "impl_latency_ms": r"\*\*实现平均延迟\*\*:\s*([\d.]+)\s*ms",
    "geo_speedup": r"\*\*几何平均加速比\*\*:\s*([\d.]+)x",
}


def parse_report_md(path: Path):
    """从 report.md 提取关键字段。找不到返回空 dict。"""
    text = path.read_text(encoding="utf-8")
    out = {}
    for key, pat in REPORT_PATTERNS.items():
        m = re.search(pat, text)
        if not m:
            continue
        if key == "pass_rate_passed":
            out["passed_cases"] = int(m.group(1))
            out["total_cases"] = int(m.group(2))
        else:
            out[key] = m.group(1)
    return out


# ── summary.json 解析 ───────────────────────────────────────────────────────
def _mean(values):
    vals = [v for v in values if v is not None]
    return sum(vals) / len(vals) if vals else None


def parse_summary_json(path: Path):
    """从 summary.json 提取关键字段，找不到返回空 dict。"""
    with open(path, encoding="utf-8") as f:
        d = json.load(f)
    perf = d.get("perf_data") or {}

    # impl latency: 顶层 avg_latency_ms
    impl_lat = perf.get("avg_latency_ms")

    # framework latency: 顶层无存储，按 per_shape_results 平均
    fw_lats = []
    for psr in perf.get("per_shape_results") or []:
        v = psr.get("framework_avg_latency_ms")
        if v is not None:
            fw_lats.append(v)
    fw_lat = _mean(fw_lats) if fw_lats else None

    return {
        "op_name": d.get("op_name"),
        "target_speedup": d.get("target_speedup"),
        "target_reached": d.get("target_reached"),
        "best_speedup": d.get("best_speedup"),
        "passed_cases": perf.get("passed_cases"),
        "total_cases": perf.get("total_cases"),
        "failed_cases": perf.get("failed_cases"),
        "framework_latency_ms": fw_lat,
        "impl_latency_ms": impl_lat,
        "geo_speedup": perf.get("speedup_vs_torch"),
        "success": d.get("success"),
    }


def parse_perf_result_json(path: Path):
    """补充延迟字段：perf_result.json 顶层有 framework/implementation.avg_latency_ms。"""
    if not path.exists():
        return {}
    with open(path, encoding="utf-8") as f:
        d = json.load(f)
    out = {}
    fw = d.get("framework") or {}
    impl = d.get("implementation") or {}
    if isinstance(fw, dict) and fw.get("avg_latency_ms") is not None:
        out["framework_latency_ms"] = fw["avg_latency_ms"]
    if isinstance(impl, dict) and impl.get("avg_latency_ms") is not None:
        out["impl_latency_ms"] = impl["avg_latency_ms"]
    if d.get("speedup_vs_torch") is not None:
        out["geo_speedup"] = d["speedup_vs_torch"]
    return out


def _iter_rank(name: str) -> int:
    """迭代目录排序权重：opt_iter_N 优于 iter_N，N 越大越新。"""
    m = re.match(r"(?:iter|opt_iter)_(\d+)", name)
    if not m:
        return -1
    n = int(m.group(1))
    return n + (1000 if name.startswith("opt_iter") else 0)


def _matching_verify_file(perf_f: Path):
    """按 perf 文件名匹配对应 verify 结果文件，缺省回退 verify_result.json。"""
    verify_dir = perf_f.parent / "verify"
    if perf_f.name.startswith("optimized_"):
        cand = verify_dir / "verify_result_optimized.json"
    elif perf_f.name.startswith("baseline_"):
        cand = verify_dir / "verify_result_baseline.json"
    else:
        cand = verify_dir / "verify_result.json"
    if cand.exists():
        return cand
    default = verify_dir / "verify_result.json"
    return default if default.exists() else None


def _load_json(path: Path):
    """读 JSON；解析失败记 warning 并返回 None（不静默吞异常）。"""
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError) as e:
        logger.warning("[WARN] 跳过无法解析的 JSON %s: %s: %s", path, type(e).__name__, e)
        return None


def _fallback_pass_counts(perf_f: Path, perf_json: dict):
    """取该迭代的 (passed, total)：优先同目录 verify 结果，缺失则回退 perf 文件自身。"""
    verify_file = _matching_verify_file(perf_f)
    if verify_file is not None:
        vj = _load_json(verify_file)
        if vj is not None:
            passed, total = vj.get("passed_cases"), vj.get("total_cases")
            if passed is not None and total is not None:
                return passed, total
    return perf_json.get("passed_cases"), perf_json.get("total_cases")


def _perf_candidate(perf_f: Path):
    """把单个 *perf_result*.json 装配成 (speedup, data)；不可用返回 None。"""
    j = _load_json(perf_f)
    if j is None:
        return None
    sp = j.get("speedup_vs_torch")
    if sp is None:
        return None

    data = {}
    fw = j.get("framework") or {}
    impl = j.get("implementation") or {}
    if isinstance(fw, dict) and fw.get("avg_latency_ms") is not None:
        data["framework_latency_ms"] = fw["avg_latency_ms"]
    if isinstance(impl, dict) and impl.get("avg_latency_ms") is not None:
        data["impl_latency_ms"] = impl["avg_latency_ms"]
    data["geo_speedup"] = sp

    passed, total = _fallback_pass_counts(perf_f, j)
    if passed is not None and total is not None:
        data["passed_cases"] = passed
        data["total_cases"] = total
    return sp, data


def _iter_dir_candidates(sub: Path):
    """单个迭代目录内所有可用候选 [(speedup, data, sub), ...]。"""
    out = []
    for perf_f in sub.glob("*perf_result*.json"):
        cand = _perf_candidate(perf_f)
        if cand is not None:
            out.append((cand[0], cand[1], sub))
    return out


def parse_output_fallback(op_dir: Path):
    """report.md/summary.json 均缺失时，扫描 output/ 下各迭代目录。

    取加速比最高的 *_perf_result*.json，精度取自同目录 verify 下对应的
    verify_result*.json（缺省回退 perf 文件自身 passed/total）。
    返回 (data_dict, source_rel_path)。
    """
    out_dir = op_dir / "output"
    candidates = []  # (speedup, data, subdir)
    if out_dir.is_dir():
        for sub in sorted(p for p in out_dir.iterdir() if p.is_dir()):
            candidates.extend(_iter_dir_candidates(sub))

    if not candidates:
        return {}, ""
    # 加速比最高；同分取更新迭代
    candidates.sort(key=lambda c: (c[0], _iter_rank(c[2].name)), reverse=True)
    _, data, sub = candidates[0]
    return data, f"output/{sub.name}/"


def parse_precheck_json(path: Path):
    if not path.exists():
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        return {}
    return {
        "category": d.get("category"),
        "op_subtype": d.get("op_subtype"),
    }


# ── 行装配 ──────────────────────────────────────────────────────────────────
def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_bool(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip() in ("是", "true", "True", "1", "yes", "Yes")
    return False


@dataclass
class ReportContext:
    """一次批跑的公共上下文（来自 batch_report.md 头部与命令行参数）。"""
    output_dir: Path
    header: dict
    arch_full: str
    level: str
    benchmark_dir: str = ""
    target_speedup: float = 2.0

    @property
    def arch_short(self) -> str:
        return self.arch_full.replace("ascend", "")

    @property
    def level_label(self) -> str:
        lv = str(self.level)
        return lv if lv.startswith("level") else f"level{lv}"


def _merge_missing(data: dict, extra: dict):
    """只补空缺字段，不覆盖已有值。"""
    for k, v in extra.items():
        if v is not None and data.get(k) in (None, ""):
            data[k] = v


def _collect_op_data(op_dir: Path):
    """按优先级合并该算子的数据来源，返回 (data, data_source)。

    report.md 优先；summary.json 次之；perf_result.json 补充延迟；
    前两者均缺失时回退扫描 output/ 各迭代目录（加速比最高版本）。
    """
    report_md = op_dir / "report.md"
    summary_json = op_dir / "summary.json"
    perf_json = op_dir / "output" / "perf_result.json"

    if not (report_md.exists() or summary_json.exists()):
        return parse_output_fallback(op_dir)

    data, data_source = {}, ""
    if report_md.exists():
        data.update(parse_report_md(report_md))
        data_source = "report.md"
    if summary_json.exists():
        _merge_missing(data, parse_summary_json(summary_json))
        if not data_source:
            data_source = "summary.json"
    if perf_json.exists():
        _merge_missing(data, parse_perf_result_json(perf_json))
    return data, data_source


def _fill_category(row: dict, op_dir: Path):
    """从 precheck.json 取算子分类（category 缺省用 op_subtype 顶上）。"""
    precheck_json = op_dir / "precheck.json"
    if not precheck_json.exists():
        return
    pc = parse_precheck_json(precheck_json)
    if pc.get("category"):
        row["category"] = pc["category"]
    if pc.get("op_subtype") and not row["category"]:
        row["category"] = pc["op_subtype"]


def _fill_metrics(row: dict, data: dict, target_speedup: float):
    """把解析出的数据填进 row 的数值与判定字段。"""
    row["ref_latency_ms"] = _to_float(data.get("framework_latency_ms"))
    row["impl_latency_ms"] = _to_float(data.get("impl_latency_ms"))

    sp = _to_float(data.get("geo_speedup")) or _to_float(data.get("best_speedup"))
    row["speedup"] = sp

    passed, total = data.get("passed_cases"), data.get("total_cases")
    if passed is not None and total is not None:
        row["precision"] = f"{passed}/{total}"
        row["precision_ok"] = (passed == total and total > 0)

    if sp is not None:
        row["perf_06"] = sp >= 0.6
        row["perf_target"] = sp >= target_speedup

    ts = _to_float(data.get("target_speedup"))
    if ts is not None:
        row["target_speedup"] = ts
    row["target_reached"] = _to_bool(data.get("target_reached"))

    bs = _to_float(data.get("best_speedup"))
    if bs is not None:
        row["best_speedup"] = bs


def _blank_row(op: dict, op_name_full: str, ctx: "ReportContext"):
    """未取到任何数据时的行骨架。"""
    return {
        "task_type": "算子生成",
        "hw": ctx.arch_short,
        "level": ctx.level_label,
        "op_id": op["id"],
        "op_name": op_name_full,
        "category": "",
        "ref_latency_ms": None,
        "impl_latency_ms": None,
        "speedup": None,
        "precision": "",
        "precision_ok": False,
        "perf_06": False,
        "perf_target": False,
        "npu": ctx.header.get("npu", ctx.header.get("npu_list", "")),
        "elapsed_s": op["elapsed"],
        "final_status": "成功" if "成功" in op["status"] else "失败",
        "target_speedup": None,
        "target_reached": False,
        "best_speedup": None,
        "data_source": "",
    }


def build_row(op: dict, ctx: "ReportContext"):
    """从 batch_report.md 一行 op 信息出发，结合 op_dir 内文件，装配 Excel 行。

    达标阈值取自 ctx.target_speedup（speedup >= 该值视为达标），用于最后一列性能 badge。
    """
    op_file = op["file"]
    op_name_full = op_file[:-3] if op_file.endswith(".py") else op_file
    op_dir = ctx.output_dir / op_name_full

    row = _blank_row(op, op_name_full, ctx)
    if not op_dir.exists():
        return row  # op 目录缺失，留空

    data, row["data_source"] = _collect_op_data(op_dir)
    _fill_category(row, op_dir)
    _fill_metrics(row, data, ctx.target_speedup)
    return row


# ── Excel 写入 ──────────────────────────────────────────────────────────────
# 最后一列性能阈值由 --target-speedup 控制（默认 2.0x），动态拼装表头。
PERF_06_HEADER = "性能0.6x pytorch"
COL_WIDTHS = [10, 8, 8, 11, 38, 14, 14, 18, 12, 10, 10, 14, 14, 8, 10, 10, 26]


def _perf_target_header(target_speedup: float) -> str:
    # 统一显示为 "性能2.0x pytorch" 形式（即使传入 2 也显示 2.0x，与 0.6x 对齐）
    return f"性能{target_speedup:.1f}x pytorch"


def _ok_badge(ok: bool) -> str:
    return "✅" if ok else "❌"


def _excel_headers(target_speedup: float):
    return [
        "任务类型", "硬件", "Level", "Problem ID", "算子名称", "分类",
        "参考延迟(ms)", "生成代码延迟(ms)", "加速比", "精度", "精度正确",
        PERF_06_HEADER, _perf_target_header(target_speedup),
        "NPU", "耗时(s)", "最终状态", "数据来源",
    ]


def _write_title_and_meta(ws, ctx: "ReportContext", n_cols: int):
    """写标题行与元信息行，返回 (表头行号, 元信息行数)。"""
    header = ctx.header
    title = (
        f"Triton算子生成Agent批跑评测结果 "
        f"({header.get('start_time', '')} → {header.get('end_time', '')})"
    )
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    npu_disp = header.get("npu_list") or header.get("npu", "")
    meta_lines = [
        f"benchmark: {header.get('benchmark', ctx.benchmark_dir)}",
        f"level: {ctx.level}    arch: {ctx.arch_full}    npu: {npu_disp}    "
        f"执行模式: {header.get('mode', '')}",
        f"开始: {header.get('start_time', '')}    结束: {header.get('end_time', '')}",
    ]
    for i, line in enumerate(meta_lines):
        r = 2 + i
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    return 2 + len(meta_lines) + 1, len(meta_lines)   # 表头前空一行


def _write_header_row(ws, headers: list, header_row: int):
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    head_font = Font(bold=True)
    head_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = head_align


def _row_values(row: dict):
    return [
        row["task_type"], row["hw"], row["level"], row["op_id"],
        row["op_name"], row["category"],
        row["ref_latency_ms"], row["impl_latency_ms"],
        f"{row['speedup']:.4f}x" if row["speedup"] is not None else "",
        row["precision"],
        _ok_badge(row["precision_ok"]),
        _ok_badge(row["perf_06"]),
        _ok_badge(row["perf_target"]),
        row["npu"], row["elapsed_s"], row["final_status"],
        row["data_source"],
    ]


def _write_data_rows(ws, rows: list, header_row: int):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(rows):
        r = header_row + 1 + i
        for c, v in enumerate(_row_values(row), start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center
            # 数值列右对齐
            if c in (7, 8) and isinstance(v, (int, float)):
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")


def _write_summary(ws, rows: list, ctx: "ReportContext", header_row: int, n_cols: int):
    """写汇总块，返回统计 dict。"""
    total = len(rows)
    succ = sum(1 for r in rows if r["final_status"] == "成功")
    p_ok = sum(1 for r in rows if r["precision_ok"])
    p06 = sum(1 for r in rows if r["perf_06"])
    p_target = sum(1 for r in rows if r["perf_target"])

    def pct(x):
        return f"{x / total * 100:.1f}%" if total else "0.0%"

    sum_start = header_row + 1 + len(rows) + 1  # 空一行
    ws.cell(row=sum_start, column=1, value="汇总").font = Font(bold=True)
    summaries = [
        f"总数: {total}",
        f"最终成功: {succ} ({pct(succ)})",
        f"精度正确: {p_ok} ({pct(p_ok)})",
        f"性能 ≥ 0.6x pytorch: {p06} ({pct(p06)})",
        f"性能 ≥ {ctx.target_speedup:.1f}x pytorch: {p_target} ({pct(p_target)})",
    ]
    for i, text in enumerate(summaries):
        r = sum_start + 1 + i
        ws.cell(row=r, column=1, value=text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    return {
        "total": total, "success": succ, "precision_ok": p_ok,
        "perf_06": p06, "perf_target": p_target,
    }


def _apply_layout(ws, header_row: int, n_meta: int):
    """列宽、冻结表头与行高。"""
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.row_dimensions[1].height = 22
    for r in range(2, 2 + n_meta):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[header_row].height = 32


def write_excel(out_path: Path, rows: list, ctx: "ReportContext"):
    """渲染 Excel 并落盘，返回统计 dict。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批跑结果"

    headers = _excel_headers(ctx.target_speedup)
    n_cols = len(headers)
    header_row, n_meta = _write_title_and_meta(ws, ctx, n_cols)
    _write_header_row(ws, headers, header_row)
    _write_data_rows(ws, rows, header_row)
    stats = _write_summary(ws, rows, ctx, header_row, n_cols)
    _apply_layout(ws, header_row, n_meta)

    wb.save(out_path)
    return stats


# ── main ────────────────────────────────────────────────────────────────────
def _resolve_out_path(output_dir: Path):
    """已存在 batch_report.xlsx 时不覆盖，改为生成带时间戳的新文件。"""
    out_path = output_dir / "batch_report.xlsx"
    if not out_path.exists():
        return out_path
    # 显式带时区取当前时间，再转本地时区用于文件名
    ts = datetime.now(timezone.utc).astimezone().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"batch_report_{ts}.xlsx"
    i = 1
    while out_path.exists():
        out_path = output_dir / f"batch_report_{ts}_{i}.xlsx"
        i += 1
    return out_path


def _build_argparser():
    ap = argparse.ArgumentParser(
        description="从批跑输出目录生成结果 Excel (batch_report.xlsx)"
    )
    ap.add_argument("--output-dir", required=True,
                    help="批跑输出目录（含 batch_report.md）")
    ap.add_argument("--benchmark-dir", default="",
                    help="benchmark 根目录（仅用于显示，缺省从 batch_report.md 读取）")
    ap.add_argument("--level", default="",
                    help="Level 编号（缺省从 batch_report.md 读取）")
    ap.add_argument("--arch", default="ascend910b2",
                    help="目标架构，默认 ascend910b2")
    ap.add_argument("--target-speedup", type=float, default=2.0,
                    help="性能达标阈值（speedup >= 该值视为达标），默认 2.0")
    return ap


def main():
    setup_logger(logger)
    args = _build_argparser().parse_args()

    output_dir = Path(args.output_dir).resolve()
    batch_report = output_dir / "batch_report.md"
    if not batch_report.exists():
        logger.error("[ERROR] 未找到 batch_report.md: %s", batch_report)
        sys.exit(1)

    header, op_rows = parse_batch_report(batch_report)
    if not op_rows:
        logger.warning("[WARN] batch_report.md 中未解析到算子行: %s", batch_report)

    ctx = ReportContext(
        output_dir=output_dir,
        header=header,
        arch_full=args.arch or header.get("arch", "ascend910b2"),
        level=args.level or header.get("level", ""),
        benchmark_dir=args.benchmark_dir,
        target_speedup=args.target_speedup,
    )
    rows = [build_row(op, ctx) for op in op_rows]

    out_path = _resolve_out_path(output_dir)
    stats = write_excel(out_path, rows, ctx)

    logger.info("[OK] Excel 生成: %s", out_path)
    logger.info("     共 %d 个算子 - 成功 %d, 精度通过 %d, 性能≥0.6x %d, 性能≥%.1fx %d",
                stats["total"], stats["success"], stats["precision_ok"],
                stats["perf_06"], args.target_speedup, stats["perf_target"])


if __name__ == "__main__":
    main()
